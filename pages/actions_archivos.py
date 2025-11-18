import os
import time
import csv
import json
import xml.etree.ElementTree as ET
from typing import Union, Optional, Dict, Any, List
import openpyxl
from zipfile import BadZipFile
from openpyxl.utils.exceptions import InvalidFileException
import pandas as pd
from playwright.sync_api import Page, Locator, expect, Error, TimeoutError

import allure

class FileActions:
    
    @allure.step("Inicializando la clase de Acciones de archivos")
    def __init__(self, base_page):
        self.base = base_page
        self.page: Page = base_page.page
        self.logger = base_page.logger
        self.registrar_paso = base_page.registrar_paso
    
    @allure.step("Cargando archivo(s) en el selector: '{selector}'")        
    def cargar_archivo(self, selector: Union[str, Locator], nombre_base: str, directorio: str, base_dir: str, file_names: Union[str, List[str]], tiempo: Union[int, float] = 0.5) -> bool:
        """
        Carga uno o varios archivos en un elemento de entrada de tipo 'file' en la página.
        Verifica que los archivos existan localmente antes de intentar cargarlos.
        Mide el rendimiento de la operación de carga de archivos.

        Args:
            selector (Union[str, Locator]): El **selector del elemento de entrada de archivo** (input[type="file"]).
                                            Puede ser una cadena (CSS, XPath, etc.) o un objeto `Locator` de Playwright.
            nombre_base (str): Nombre base utilizado para las **capturas de pantalla**
                               tomadas durante la ejecución de la función.
            directorio (str): **Ruta del directorio** donde se guardarán las capturas de pantalla.
            base_dir (str): **Directorio base** donde se encuentran los archivos a cargar.
            file_names (Union[str, List[str]]): El **nombre o una lista de nombres de archivo(s)**
                                                  (solo el nombre del archivo, no la ruta completa)
                                                  que se desea cargar. Estos nombres se combinarán
                                                  con `base_dir` para obtener la ruta completa.
            tiempo (Union[int, float]): **Tiempo máximo de espera** (en segundos) para que el
                                        elemento esté visible y habilitado. También es el tiempo
                                        de espera fijo después de la operación exitosa.
                                        Por defecto, `5.0` segundos (ajustado para robustez).

        Returns:
            bool: `True` si el archivo(s) se carga(n) exitosamente; `False` en caso de fallo
                  (ej., archivo no encontrado, timeout, elemento no interactuable).

        Raises:
            FileNotFoundError: Si alguno de los archivos especificados no existe en el `base_dir`.
            Error: Si ocurre un problema específico de Playwright (ej., selector inválido,
                   elemento no es un input de tipo file, timeout de visibilidad/habilitación).
            Exception: Para cualquier otro error inesperado.
        """
        # Normalizar `file_names` a una lista para manejar consistentemente uno o varios archivos
        file_names_list = [file_names] if isinstance(file_names, str) else file_names
        
        files_str = ', '.join(file_names_list)
        nombre_paso = f"Cargar archivo(s): '{files_str}' en selector: '{selector}' (Base: '{base_dir}')"
        self.registrar_paso(nombre_paso)

        self.logger.info(f"\nIntentando cargar archivo(s) '{file_names_list}' en el selector: '{selector}'. Tiempo máximo de espera: {tiempo}s.")

        # Construir las rutas completas de los archivos y verificar su existencia localmente
        full_file_paths = []
        for name in file_names_list:
            full_path = os.path.join(base_dir, name)
            full_file_paths.append(full_path)
            self.logger.debug(f"\nConstruida ruta completa para archivo: '{full_path}'")

            if not os.path.exists(full_path):
                error_msg = f"\n❌ Error: El archivo no existe en la ruta especificada: '{full_path}'."
                self.logger.error(error_msg, exc_info=True)
                self.base.tomar_captura(f"{nombre_base}_archivo_no_encontrado", directorio)
                raise FileNotFoundError(error_msg) # Elevar un error específico si el archivo no se encuentra.

        # Asegura que 'selector' sea un objeto Locator de Playwright para un uso consistente.
        if isinstance(selector, str):
            locator = self.page.locator(selector)
        else:
            locator = selector

        # --- Medición de rendimiento: Inicio de la operación de carga de archivos ---
        # Registra el tiempo justo antes de iniciar la interacción con el elemento de entrada de archivo.
        start_time_file_upload = time.time()

        try:
            # 1. Esperar a que el elemento de entrada de archivo esté visible y habilitado
            # Es fundamental asegurar que el elemento está listo para interactuar.
            self.logger.debug(f"\nEsperando que el selector '{selector}' esté visible y habilitado (timeout: {tiempo}s).")
            expect(locator).to_be_visible()
            expect(locator).to_be_enabled() # También se puede usar to_be_editable() si es un input
            self.logger.info(f"\nEl selector '{selector}' está visible y habilitado.")

            # 2. Opcional: Resaltar el elemento para depuración visual
            locator.highlight()
            self.logger.debug(f"\nElemento con selector '{selector}' resaltado.")
            self.base.tomar_captura(f"{nombre_base}_antes_cargar_archivos", directorio) # Captura antes de adjuntar los archivos.

            # 3. Usar set_input_files para adjuntar el archivo(s)
            # Playwright maneja la interacción con el diálogo de carga de archivos.
            # Se le pasa una lista de rutas completas de los archivos a adjuntar.
            self.logger.info(f"\nAdjuntando archivo(s) {file_names_list} al selector '{selector}'.")
            locator.set_input_files(full_file_paths)

            # --- Medición de rendimiento: Fin de la operación de carga de archivos ---
            # Registra el tiempo una vez que Playwright ha adjuntado los archivos.
            end_time_file_upload = time.time()
            duration_file_upload = end_time_file_upload - start_time_file_upload
            self.logger.info(f"PERFORMANCE: Tiempo que tardó en cargar el archivo(s) '{file_names_list}' en el selector '{selector}': {duration_file_upload:.4f} segundos.")

            # Construir mensaje de éxito basado en si es uno o varios archivos
            if len(file_names_list) == 1:
                success_msg = f"\n✅ Archivo '{file_names_list[0]}' cargado exitosamente desde '{base_dir}' en el selector '{selector}'."
            else:
                success_msg = f"\n✅ Archivos {file_names_list} cargados exitosamente desde '{base_dir}' en el selector '{selector}'."
            self.logger.info(success_msg)
            
            self.base.tomar_captura(f"{nombre_base}_archivos_cargados", directorio)
            return True

        except TimeoutError as e:
            # Captura si el elemento no se hace visible o habilitado a tiempo.
            error_files_info = file_names_list[0] if len(file_names_list) == 1 else file_names_list
            end_time_fail = time.time()
            duration_fail = end_time_fail - start_time_file_upload # Mide desde el inicio de la operación.
            error_msg = (
                f"\n❌ FALLO (Timeout): El elemento '{selector}' no estuvo visible o habilitado "
                f"después de {duration_fail:.4f} segundos (timeout configurado: {tiempo}s) para cargar el archivo(s) '{error_files_info}'. "
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True) # Usa 'error' porque un timeout al cargar archivos es un fallo crítico.
            self.base.tomar_captura(f"{nombre_base}_fallo_timeout_cargar_archivo", directorio)
            return False

        except Error as e:
            # Captura errores específicos de Playwright (ej., selector inválido, el elemento no es un input[type="file"]).
            error_files_info = file_names_list[0] if len(file_names_list) == 1 else file_names_list
            error_msg = (
                f"\n❌ FALLO (Playwright): Error de Playwright al cargar el archivo(s) '{error_files_info}' "
                f"en el selector '{selector}'. Esto puede deberse a un selector incorrecto o que el elemento "
                f"no es un input de tipo archivo válido.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.base.tomar_captura(f"{nombre_base}_error_playwright_cargar_archivo", directorio)
            raise # Re-lanza la excepción porque es un fallo de ejecución.

        except Exception as e:
            # Captura cualquier otra excepción inesperada.
            error_files_info = file_names_list[0] if len(file_names_list) == 1 else file_names_list
            error_msg = (
                f"\n❌ FALLO (Inesperado): Ocurrió un error inesperado al cargar el archivo(s) '{error_files_info}' "
                f"en el selector '{selector}'.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.base.tomar_captura(f"{nombre_base}_error_inesperado_cargar_archivo", directorio)
            raise # Re-lanza la excepción.
        
    @allure.step("Removiendo carga de archivo en el selector: '{selector}'")
    def remover_carga_de_archivo(self, selector: Union[str, Locator], nombre_base: str, directorio: str, tiempo: Union[int, float] = 0.5) -> bool:
        """
        Remueve la carga de archivo(s) de un elemento de entrada de tipo 'file'
        estableciendo su valor a una lista vacía. Mide el rendimiento de esta operación.

        Args:
            selector (Union[str, Locator]): El **selector del elemento de entrada de archivo** (input[type="file"])
                                            del cual se removerá la carga.
                                            Puede ser una cadena (CSS, XPath, etc.) o un objeto `Locator` de Playwright.
            nombre_base (str): Nombre base utilizado para las **capturas de pantalla**
                               tomadas durante la ejecución de la función.
            directorio (str): **Ruta del directorio** donde se guardarán las capturas de pantalla.
            tiempo (Union[int, float]): **Tiempo máximo de espera** (en segundos) para que el
                                        elemento esté visible y habilitado antes de intentar
                                        remover la carga. También es el tiempo de espera fijo
                                        después de la operación exitosa.
                                        Por defecto, `5.0` segundos (ajustado para robustez).

        Returns:
            bool: `True` si la carga del archivo se remueve exitosamente; `False` en caso de fallo
                  (ej., timeout, elemento no interactuable).

        Raises:
            Error: Si ocurre un problema específico de Playwright (ej., selector inválido,
                   elemento no es un input de tipo file, timeout de visibilidad/habilitación).
            Exception: Para cualquier otro error inesperado.
        """
        nombre_paso = f"Remover carga de archivo del input: '{selector}'"
        self.registrar_paso(nombre_paso)
        
        self.logger.info(f"\nIntentando remover la carga de archivo para el selector: '{selector}'. Tiempo máximo de espera: {tiempo}s.")

        # Asegura que 'selector' sea un objeto Locator de Playwright para un uso consistente.
        if isinstance(selector, str):
            locator = self.page.locator(selector)
        else:
            locator = selector

        # --- Medición de rendimiento: Inicio de la operación de remoción de archivos ---
        # Registra el tiempo justo antes de iniciar la interacción con el elemento.
        start_time_file_removal = time.time()

        try:
            # 1. Esperar a que el elemento de entrada de archivo esté visible y habilitado
            # Es fundamental asegurar que el elemento está listo para interactuar y aceptar la limpieza.
            self.logger.debug(f"\nEsperando que el selector '{selector}' esté visible y habilitado (timeout: {tiempo}s) para remover la carga.")
            expect(locator).to_be_visible()
            expect(locator).to_be_enabled() # O to_be_editable()
            self.logger.info(f"\nEl selector '{selector}' está visible y habilitado.")

            # 2. Resaltar el elemento para depuración visual
            locator.highlight()
            self.logger.debug(f"\nElemento con selector '{selector}' resaltado.")
            self.base.tomar_captura(f"{nombre_base}_antes_remover_carga", directorio) # Captura antes de remover.

            # 3. Usar set_input_files con una lista vacía para remover el archivo
            # Esto simula el usuario cancelando o limpiando la selección de archivos.
            self.logger.info(f"\nEstableciendo input files a vacío para el selector '{selector}'.")
            locator.set_input_files([])

            # --- Medición de rendimiento: Fin de la operación de remoción de archivos ---
            # Registra el tiempo una vez que Playwright ha limpiado el input de archivos.
            end_time_file_removal = time.time()
            duration_file_removal = end_time_file_removal - start_time_file_removal
            self.logger.info(f"PERFORMANCE: Tiempo que tardó en remover la carga de archivo para el selector '{selector}': {duration_file_removal:.4f} segundos.")

            self.logger.info(f"\n✅ Carga de archivo removida exitosamente para el selector '{selector}'.")
            self.base.tomar_captura(f"{nombre_base}_remocion_completa", directorio)
            return True

        except TimeoutError as e:
            # Captura si el elemento no se hace visible o habilitado a tiempo.
            end_time_fail = time.time()
            duration_fail = end_time_fail - start_time_file_removal # Mide desde el inicio de la operación.
            error_msg = (
                f"\n❌ FALLO (Timeout): El elemento '{selector}' no estuvo visible o habilitado "
                f"después de {duration_fail:.4f} segundos (timeout configurado: {tiempo}s) para remover la carga de archivo. "
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True) # Usa 'error' porque un timeout es un fallo crítico.
            self.base.tomar_captura(f"{nombre_base}_fallo_timeout_remocion_archivo", directorio)
            return False

        except Error as e:
            # Captura errores específicos de Playwright (ej., selector inválido, el elemento no es un input[type="file"]).
            error_msg = (
                f"\n❌ FALLO (Playwright): Error de Playwright al intentar remover la carga de archivo "
                f"para el selector '{selector}'. Esto puede deberse a un selector incorrecto o que el elemento "
                f"no es un input de tipo archivo válido.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.base.tomar_captura(f"{nombre_base}_error_playwright_remocion_archivo", directorio)
            raise # Re-lanza la excepción porque es un fallo de ejecución.

        except Exception as e:
            # Captura cualquier otra excepción inesperada.
            error_msg = (
                f"\n❌ FALLO (Inesperado): Ocurrió un error inesperado al intentar remover la carga de archivo "
                f"para el selector '{selector}'.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.base.tomar_captura(f"{nombre_base}_error_inesperado_remocion_archivo", directorio)
            raise # Re-lanza la excepción.
                
    @allure.step("Descargando archivo desde el selector: '{selector}', en el directorio: '{directorio_descargas}'")
    def descargar_archivo(self, selector: Union[str, Locator], nombre_base: str, directorio_capturas: str, directorio_descargas: str, tiempo: Union[int, float] = 30.0) -> str:
        """
        Descarga un archivo al hacer clic en un selector específico.
        
        Esta función espera que se inicie una descarga, la guarda en un directorio local
        especificado y mide el tiempo de la operación. Es ideal para elementos como enlaces
        o botones que inician la descarga de un archivo.

        Args:
            selector (Union[str, Locator]): El selector del elemento (enlace, botón, etc.) 
                                            que desencadena la descarga.
            nombre_base (str): Nombre base para las capturas de pantalla tomadas durante la ejecución.
            directorio_capturas (str): Ruta del directorio donde se guardarán las capturas de pantalla.
            directorio_descargas (str): El directorio de destino donde se guardará el archivo descargado.
            tiempo (Union[int, float], opcional): Tiempo máximo de espera (en segundos) para que la 
                                                descarga se complete. Por defecto, 30.0 segundos.
                                                
        Returns:
            str: La ruta completa del archivo descargado si la operación es exitosa; 
                `None` en caso de cualquier fallo.

        Raises:
            TimeoutError: Si la descarga no se inicia o no se completa dentro del tiempo especificado.
            Error: Si ocurre un problema específico de Playwright, como un selector no válido.
            Exception: Para cualquier otro error inesperado.
        """
        nombre_paso = f"Descargando archivo desde el selector: '{selector}', en el directorio: '{directorio_descargas}'"
        self.registrar_paso(nombre_paso)
        
        # 1. Asegurar que el selector sea un objeto Locator para un uso uniforme.
        locator = self.page.locator(selector) if isinstance(selector, str) else selector
        self.logger.info(f"\nIntentando descargar archivo desde el selector: '{selector}'. Tiempo máximo de espera: {tiempo}s.")

        # 2. Configurar la escucha de la descarga ANTES de la acción que la desencadena.
        #    La declaración `with` asegura que la escucha se active antes de hacer clic.
        start_time_download = time.time()
        try:
            with self.page.expect_download() as download_info:
                # 3. Realizar la acción que inicia la descarga (ej. hacer clic en un enlace).
                self.logger.info(f"\nRealizando la acción de clic en el selector '{selector}' para iniciar la descarga.")
                locator.click()

            # 4. Obtener el objeto de descarga y la ruta temporal del archivo.
            download = download_info.value
            path_temp = download.path()
            file_name = download.suggested_filename

            # 5. Definir la ruta de destino y mover el archivo descargado.
            #    Es crucial mover el archivo desde su ruta temporal antes de que Playwright
            #    limpie la sesión.
            ruta_completa_del_archivo = os.path.join(directorio_descargas, file_name)
            os.makedirs(directorio_descargas, exist_ok=True) # Crea el directorio si no existe.
            download.save_as(ruta_completa_del_archivo)
            self.logger.info(f"\nArchivo guardado exitosamente: '{ruta_completa_del_archivo}'.")

            # 6. Medición de rendimiento y registro de éxito.
            end_time_download = time.time()
            duration_download = end_time_download - start_time_download
            self.logger.info(f"PERFORMANCE: Tiempo que tardó en descargar el archivo '{file_name}': {duration_download:.4f} segundos.")
            self.logger.info(f"\n✅ Archivo descargado exitosamente y guardado en '{ruta_completa_del_archivo}'.")
            self.base.tomar_captura(f"{nombre_base}_archivo_descargado", directorio_capturas)
            return ruta_completa_del_archivo

        except TimeoutError as e:
            # Manejo de error: la descarga no se inició o no se completó a tiempo.
            end_time_fail = time.time()
            duration_fail = end_time_fail - start_time_download
            error_msg = (
                f"\n❌ FALLO (Timeout): El elemento '{selector}' no estuvo visible/habilitado o "
                f"la descarga no se inició/completó después de {duration_fail:.4f} segundos.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.base.tomar_captura(f"{nombre_base}_fallo_timeout_descargar_archivo", directorio_capturas)
            return None

        except Error as e:
            # Manejo de error: problemas con el selector o interacción de Playwright.
            error_msg = (
                f"\n❌ FALLO (Playwright): Error de Playwright al intentar descargar "
                f"el archivo desde el selector '{selector}'.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.base.tomar_captura(f"{nombre_base}_error_playwright_descarga", directorio_capturas)
            raise # Re-lanzar la excepción para que el test falle.

        except Exception as e:
            # Manejo de cualquier otro error inesperado.
            error_msg = (
                f"\n❌ FALLO (Inesperado): Ocurrió un error inesperado al intentar descargar "
                f"el archivo desde el selector '{selector}'.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.base.tomar_captura(f"{nombre_base}_error_inesperado_descarga", directorio_capturas)
            raise # Re-lanzar la excepción.
        
    @allure.step("Obteniendo número de filas en el archivo Excel: '{archivo_excel_path}', hoja: '{hoja}', con encabezado: {has_header}")
    def num_Filas_excel(self, archivo_excel_path: str, hoja: str, has_header: bool = False, nombre_paso: str = "") -> int:
        """
        Detecta y devuelve el número total de filas ocupadas en una hoja específica de un archivo Excel.
        Opcionalmente, descuenta una fila para el encabezado si 'has_header' es True.
        Esta función mide el tiempo que tarda en cargar el archivo Excel y obtener el número de filas,
        lo cual es útil para pruebas de rendimiento en escenarios de procesamiento de datos.

        Args:
            archivo_excel_path (str): La **ruta completa al archivo Excel** (`.xlsx` o `.xlsm`).
            hoja (str): El **nombre de la hoja/pestaña** dentro del archivo Excel de la cual se desean contar las filas.
            has_header (bool, opcional): Si es `True`, se descuenta una fila del total
                                         para considerar que la primera fila es un encabezado.
                                         Por defecto es `False`.
            nombre_paso (str, opcional): Una descripción del paso que se está ejecutando para los logs. Por defecto "".

        Returns:
            int: El **número de filas de datos** en la hoja especificada.
                 Retorna `0` si el archivo no se encuentra, la hoja no existe, o si ocurre un error inesperado.
        """
        nombre_paso = f"Obteniendo número de filas en el archivo Excel: '{archivo_excel_path}', hoja: '{hoja}', con encabezado: {has_header}"
        self.registrar_paso(nombre_paso)
        
        self.logger.info(f"\n--- {nombre_paso}: Intentando obtener el número de filas para la hoja '{hoja}' en el archivo '{archivo_excel_path}' (tiene encabezado: {has_header}). ---")

        # --- Medición de rendimiento: Inicio total de la función ---
        start_time_total_operation = time.time()

        num_physical_rows = 0
        num_data_rows = 0

        try:
            self.logger.info(f"\n⏳ Cargando el libro de trabajo Excel: '{archivo_excel_path}'...")
            workbook = openpyxl.load_workbook(archivo_excel_path) # Carga el libro de trabajo Excel
            self.logger.info(f"\n✅ Libro de trabajo cargado. Seleccionando la hoja '{hoja}'...")
            sheet = workbook[hoja] # Selecciona la hoja específica del libro
            
            # Obtiene el número total de filas con contenido.
            # openpyxl.worksheet.max_row devuelve el índice de la última fila no vacía.
            num_physical_rows = sheet.max_row 

            if has_header and num_physical_rows > 0:
                # Si tiene encabezado y hay al menos una fila (el encabezado)
                num_data_rows = num_physical_rows - 1 # Resta 1 para no contar el encabezado
                self.logger.info(f"\n✅ Se encontraron {num_data_rows} filas de datos (descontando encabezado) en la hoja '{hoja}'.")
                return num_data_rows
            else:
                # Para hojas sin encabezado, o si num_physical_rows es 0 (hoja vacía),
                # el número de filas de datos es igual al número de filas físicas.
                num_data_rows = num_physical_rows
                self.logger.info(f"\n✅ Se encontraron {num_data_rows} filas ocupadas en la hoja '{hoja}'.")
                return num_data_rows

        except FileNotFoundError:
            error_msg = f"\n❌ FALLO (Archivo no encontrado): El archivo Excel no se encontró en la ruta: '{archivo_excel_path}'."
            self.logger.critical(error_msg)
            return 0
        except KeyError:
            error_msg = f"\n❌ FALLO (Hoja no encontrada): La hoja '{hoja}' no se encontró en el archivo Excel: '{archivo_excel_path}'."
            self.logger.critical(error_msg)
            return 0
        except Exception as e:
            error_msg = (
                f"\n❌ FALLO (Error Inesperado): Ocurrió un error inesperado al leer el número de filas del Excel.\n"
                f"Archivo: '{archivo_excel_path}', Hoja: '{hoja}'.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True) # Incluye el stack trace
            return 0
        finally:
            # --- Medición de rendimiento: Fin total de la función ---
            end_time_total_operation = time.time()
            duration_total_operation = end_time_total_operation - start_time_total_operation
            self.logger.info(f"PERFORMANCE: Tiempo total de la operación (num_Filas_excel): {duration_total_operation:.4f} segundos.")
            # Es importante cerrar el workbook si se ha abierto explícitamente y no con 'with open()'
            # Sin embargo, openpyxl.load_workbook no requiere un cierre explícito en la mayoría de los casos
            # ya que maneja el archivo internamente. Aun así, se puede añadir un log de depuración.
            self.logger.debug("\nFinalizada la operación de lectura de Excel.")

    @allure.step("Obteniendo dato de la celda en el archivo Excel: '{archivo_excel_path}', hoja: '{hoja}', fila lógica: {numero_fila_logica}, columna: '{nombre_o_indice_columna}', con encabezado: {has_header_excel}")
    def dato_Columna_excel(self, archivo_excel_path: str, hoja: str, numero_fila_logica: int, nombre_o_indice_columna: Union[str, int], has_header_excel: bool = False, nombre_paso: str = "") -> Union[str, int, float, None]:
        """
        Obtiene el valor de una celda específica de una hoja de un archivo Excel.
        Ajusta el número de fila si se indica que la hoja tiene un encabezado.
        Permite especificar la columna por su nombre (si hay encabezado) o por su índice numérico.
        Esta función mide el tiempo que tarda en cargar el archivo, ubicar la columna/fila,
        y extraer el dato, lo cual es útil para identificar cuellos de botella en la lectura de datos.

        Args:
            archivo_excel_path (str): La **ruta completa al archivo Excel** (`.xlsx` o `.xlsm`).
            hoja (str): El **nombre de la hoja/pestaña** dentro del archivo Excel.
            numero_fila_logica (int): El **número de fila lógica** (basado en 1) de la celda a leer.
                                     Si `has_header_excel` es `True`, esta es la fila de datos
                                     (e.g., `1` para la primera fila después del encabezado).
            nombre_o_indice_columna (Union[str, int]): El **nombre del encabezado de la columna** (string)
                                                       o el **índice numérico de la columna** (entero, basado en 1).
            has_header_excel (bool, opcional): Si es `True`, indica que la hoja tiene un encabezado en la primera fila.
                                             Esto ajusta el cálculo de la fila física y permite la búsqueda por nombre de columna.
                                             Por defecto es `False`.
            nombre_paso (str, opcional): Una descripción del paso que se está ejecutando para los logs. Por defecto "".

        Returns:
            Union[str, int, float, None]: El valor de la celda. El tipo del valor se conserva,
                                          pero si se convierte a `str` para consumo general.
                                          Retorna `None` si el archivo no se encuentra, la hoja/columna no existe,
                                          la fila/columna está fuera de rango, o si ocurre un error.
        """
        nombre_paso = f"Obteniendo dato de la celda en el archivo Excel: '{archivo_excel_path}', hoja: '{hoja}', fila lógica: {numero_fila_logica}, columna: '{nombre_o_indice_columna}', con encabezado: {has_header_excel}"
        self.registrar_paso(nombre_paso)
        
        self.logger.info(f"\n--- {nombre_paso}: Intentando obtener dato de la celda (Fila lógica: {numero_fila_logica}, Columna: {nombre_o_indice_columna}) de la hoja '{hoja}' en el archivo '{archivo_excel_path}' (tiene encabezado: {has_header_excel}). ---")

        # --- Medición de rendimiento: Inicio total de la función ---
        start_time_total_operation = time.time()
        cell_value: Any = None # Inicializamos el valor de la celda

        try:
            # --- Medición de rendimiento: Carga del Workbook y selección de hoja ---
            start_time_load_workbook = time.time()
            self.logger.info(f"\n⏳ Cargando el libro de trabajo Excel: '{archivo_excel_path}'...")
            workbook = openpyxl.load_workbook(archivo_excel_path)
            self.logger.info(f"\n✅ Libro de trabajo cargado. Seleccionando la hoja '{hoja}'...")
            sheet = workbook[hoja]
            end_time_load_workbook = time.time()
            duration_load_workbook = end_time_load_workbook - start_time_load_workbook
            self.logger.info(f"PERFORMANCE: Tiempo de carga del workbook y selección de hoja: {duration_load_workbook:.4f} segundos.")

            # 1. Determinar el índice físico de la columna
            col_index: int = -1
            if isinstance(nombre_o_indice_columna, str):
                # --- Medición de rendimiento: Búsqueda de columna por nombre ---
                start_time_find_column = time.time()
                self.logger.info(f"\n🔎 Buscando columna por nombre: '{nombre_o_indice_columna}' en el encabezado de la hoja '{hoja}'...")
                header_found = False
                # sheet[1] se refiere a la primera fila física del Excel
                for col_idx, cell in enumerate(sheet[1], 1):
                    if cell.value is not None and str(cell.value).strip().lower() == nombre_o_indice_columna.strip().lower():
                        col_index = col_idx
                        header_found = True
                        break
                end_time_find_column = time.time()
                duration_find_column = end_time_find_column - start_time_find_column
                self.logger.info(f"PERFORMANCE: Tiempo de búsqueda de columna por nombre: {duration_find_column:.4f} segundos.")

                if not header_found:
                    self.logger.error(f"\n❌ Error: La columna '{nombre_o_indice_columna}' no fue encontrada en el encabezado de la hoja '{hoja}'.")
                    return None
            elif isinstance(nombre_o_indice_columna, int):
                col_index = nombre_o_indice_columna
            else:
                self.logger.error(f"\n❌ Error: El parámetro 'nombre_o_indice_columna' debe ser un string (nombre) o un entero (índice). Se recibió: '{nombre_o_indice_columna}' ({type(nombre_o_indice_columna).__name__}).")
                return None

            # Validar que el índice de columna sea válido
            if not (1 <= col_index <= sheet.max_column):
                self.logger.error(f"\n❌ Error: Índice de columna '{col_index}' fuera de rango para la hoja '{hoja}' (máximo: {sheet.max_column}).")
                return None

            # 2. Determinar el índice físico de la fila
            # 'numero_fila_logica' es la fila de datos que el usuario piensa (1 para la primera fila de datos).
            # Si hay encabezado, la primera fila de datos (lógica 1) está en la fila física 2.
            # Por lo tanto, sumamos 1 si hay encabezado.
            actual_fila_fisica = numero_fila_logica + 1 if has_header_excel else numero_fila_logica

            # Validar que la fila física sea válida
            if not (1 <= actual_fila_fisica <= sheet.max_row):
                self.logger.warning(f"\n⚠️ Advertencia: La fila física {actual_fila_fisica} (lógica: {numero_fila_logica}) está fuera del rango de filas de la hoja '{hoja}' (máximo: {sheet.max_row}). Retornando None.")
                return None
            
            self.logger.info(f"\n🔎 Intentando obtener el dato de la celda (Fila lógica: {numero_fila_logica}, Fila física: {actual_fila_fisica}, Columna: {nombre_o_indice_columna}) de la hoja '{hoja}'.")
            
            # --- Medición de rendimiento: Lectura de la celda ---
            start_time_read_cell = time.time()
            cell_value = sheet.cell(row=actual_fila_fisica, column=col_index).value
            end_time_read_cell = time.time()
            duration_read_cell = end_time_read_cell - start_time_read_cell
            self.logger.info(f"PERFORMANCE: Tiempo de lectura de la celda: {duration_read_cell:.4f} segundos.")
            
            # Convertir a string para asegurar que 'rellenar_campo_de_texto' u otras funciones siempre reciban un str
            if cell_value is not None:
                valor_retorno = str(cell_value)
                self.logger.info(f"\n✅ Dato obtenido de (Fila lógica: {numero_fila_logica}, Columna: {nombre_o_indice_columna}) en '{hoja}': '{valor_retorno}'.")
                return valor_retorno
            else:
                self.logger.warning(f"\n⚠️ La celda en Fila lógica: {numero_fila_logica}, Columna: {nombre_o_indice_columna} en '{hoja}' está vacía. Retornando None.")
                return None

        except FileNotFoundError:
            error_msg = f"\n❌ FALLO (Archivo no encontrado): El archivo Excel no se encontró en la ruta: '{archivo_excel_path}'."
            self.logger.critical(error_msg)
            return None
        except KeyError:
            error_msg = f"\n❌ FALLO (Hoja no encontrada): La hoja '{hoja}' no se encontró en el archivo Excel: '{archivo_excel_path}'."
            self.logger.critical(error_msg)
            return None
        except Exception as e:
            error_msg = (
                f"\n❌ FALLO (Error Inesperado): Ocurrió un error inesperado al leer el dato del Excel.\n"
                f"Archivo: '{archivo_excel_path}', Hoja: '{hoja}'.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True) # Incluye el stack trace
            return None
        finally:
            # --- Medición de rendimiento: Fin total de la función ---
            end_time_total_operation = time.time()
            duration_total_operation = end_time_total_operation - start_time_total_operation
            self.logger.info(f"PERFORMANCE: Tiempo total de la operación (dato_Columna_excel): {duration_total_operation:.4f} segundos.")
            # Aunque openpyxl maneja la liberación de recursos, un log final es útil.
            self.logger.debug("\nFinalizada la operación de lectura de dato de Excel.")
    
    @allure.step("Obteniendo número de filas en el archivo CSV: '{archivo_csv_path}', con delimitador: '{delimiter}', con encabezado: {has_header}")
    def num_Filas_csv(self, archivo_csv_path: str, delimiter: str = ',', has_header: bool = False, nombre_paso: str = "") -> int:
        """
        Detecta y devuelve el número total de filas de datos en un archivo CSV.
        Opcionalmente, descuenta una fila para el encabezado si 'has_header' es True.
        Esta función mide el tiempo que tarda en abrir el archivo CSV, leer todas sus filas
        y realizar el conteo, lo cual es útil para evaluar el rendimiento en escenarios
        de procesamiento de grandes volúmenes de datos CSV.

        Args:
            archivo_csv_path (str): La **ruta completa al archivo CSV**.
            delimiter (str, opcional): El **carácter utilizado como separador** de datos en el CSV
                                      (e.g., ',', ';', '\t'). Por defecto es `,`.
            has_header (bool, opcional): Si es `True`, se descuenta una fila del total
                                         para considerar que la primera fila es un encabezado.
                                         Por defecto es `False`.
            nombre_paso (str, opcional): Una descripción del paso que se está ejecutando para los logs. Por defecto "".

        Returns:
            int: El **número de filas de datos** en el archivo CSV.
                 Retorna `0` si el archivo no se encuentra, ocurre un error de formato CSV,
                 o si hay un error inesperado.
        """
        nombre_paso = f"Obteniendo número de filas en el archivo CSV: '{archivo_csv_path}', con delimitador: '{delimiter}', con encabezado: {has_header}"
        self.registrar_paso(nombre_paso)
        
        self.logger.info(f"\n--- {nombre_paso}: Intentando obtener el número de filas para el archivo CSV '{archivo_csv_path}' con delimitador '{delimiter}' (tiene encabezado: {has_header}). ---")

        # --- Medición de rendimiento: Inicio total de la función ---
        start_time_total_operation = time.time()
        
        row_count = 0 # Inicializamos el contador de filas

        try:
            self.logger.info(f"\n⏳ Abriendo y leyendo el archivo CSV: '{archivo_csv_path}'...")
            with open(archivo_csv_path, 'r', newline='', encoding='utf-8-sig') as csvfile:
                # Crea un objeto reader para iterar sobre las líneas del CSV, usando el delimitador especificado.
                # 'newline=''' es crucial para evitar problemas con saltos de línea en diferentes SO.
                # 'encoding='utf-8'' es una buena práctica para manejar caracteres especiales.
                csv_reader = csv.reader(csvfile, delimiter=delimiter)
                
                # Cuenta todas las filas en el CSV. sum(1 for row in csv_reader) es una forma eficiente.
                row_count = sum(1 for row in csv_reader)

            self.logger.info(f"\n✅ Lectura de archivo CSV completada. Filas totales encontradas: {row_count}.")

            if has_header and row_count > 0:
                # Si tiene encabezado y el archivo no está vacío (es decir, hay al menos el encabezado)
                num_data_rows = row_count - 1 # Resta 1 para no contar el encabezado, obteniendo solo las filas de datos
                self.logger.info(f"\n✅ Se encontraron {num_data_rows} filas de datos (descontando encabezado) en el archivo CSV '{archivo_csv_path}'.")
                return num_data_rows
            else:
                # Si no tiene encabezado o el archivo está vacío (row_count es 0 o 1 si solo es un encabezado sin datos)
                num_data_rows = row_count
                self.logger.info(f"\n✅ Se encontraron {num_data_rows} filas ocupadas en el archivo CSV '{archivo_csv_path}'.")
                return num_data_rows

        except FileNotFoundError:
            error_msg = f"\n❌ FALLO (Archivo no encontrado): El archivo CSV no se encontró en la ruta: '{archivo_csv_path}'."
            self.logger.critical(error_msg)
            return 0
        except csv.Error as e:
            error_msg = f"\n❌ FALLO (Error de formato CSV): Ocurrió un error al procesar el archivo CSV '{archivo_csv_path}'.\nDetalles: {e}"
            self.logger.critical(error_msg, exc_info=True) # Incluye el stack trace para errores de CSV
            return 0
        except Exception as e:
            error_msg = (
                f"\n❌ FALLO (Error Inesperado): Ocurrió un error desconocido al leer el número de filas del CSV.\n"
                f"Archivo: '{archivo_csv_path}'.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True) # Incluye el stack trace para errores inesperados
            return 0
        finally:
            # --- Medición de rendimiento: Fin total de la función ---
            end_time_total_operation = time.time()
            duration_total_operation = end_time_total_operation - start_time_total_operation
            self.logger.info(f"PERFORMANCE: Tiempo total de la operación (num_Filas_csv): {duration_total_operation:.4f} segundos.")
            self.logger.debug("\nFinalizada la operación de lectura de CSV.")

    @allure.step("Obteniendo dato de la celda en el archivo CSV: '{archivo_csv_path}', fila lógica: {fila_logica}, columna lógica: {columna_logica}, con delimitador: '{delimiter}', con encabezado: {has_header}")
    def dato_Columna_csv(self, archivo_csv_path: str, fila_logica: int, columna_logica: int, delimiter: str = ',', has_header: bool = False, nombre_paso: str = "") -> Optional[str]:
        """
        Obtiene el valor de una "celda" específica de un archivo CSV, ajustando el índice de la fila
        si se indica que la primera fila es un encabezado. Permite especificar el delimitador del CSV.
        Esta función mide el tiempo que tarda en cargar el archivo CSV, leer todas sus filas
        y extraer el dato de la celda solicitada, lo cual es crucial para evaluar el rendimiento
        en escenarios de automatización basados en datos de archivos CSV.

        Args:
            archivo_csv_path (str): La **ruta completa al archivo CSV**.
            fila_logica (int): El **número de fila lógico** (basado en 1) de la celda a leer.
                               Si `has_header` es `True`, esta es la fila de datos
                               (e.g., `1` para la primera fila después del encabezado).
            columna_logica (int): El **número de columna lógico** (basado en 1) de la celda a leer.
            delimiter (str, opcional): El **carácter utilizado como separador** de datos en el CSV
                                      (e.g., ',', ';', '\t'). Por defecto es `,`.
            has_header (bool, opcional): Si es `True`, indica que la primera fila del CSV es un encabezado.
                                         Esto ajusta el cálculo de la fila física. Por defecto es `False`.
            nombre_paso (str, opcional): Una descripción del paso que se está ejecutando para los logs. Por defecto "".

        Returns:
            Optional[str]: El **valor de la celda como string**. Retorna `None` si el archivo no se encuentra,
                           los índices de fila/columna están fuera de rango, hay un error de formato CSV,
                           o si ocurre un error inesperado.
        """
        nombre_paso = f"Obteniendo dato de la celda en el archivo CSV: '{archivo_csv_path}', fila lógica: {fila_logica}, columna lógica: {columna_logica}, con delimitador: '{delimiter}', con encabezado: {has_header}"
        self.registrar_paso(nombre_paso)
        
        self.logger.info(f"\n--- {nombre_paso}: Intentando obtener dato de la celda (Fila lógica: {fila_logica}, Columna lógica: {columna_logica}) del archivo CSV '{archivo_csv_path}' con delimitador '{delimiter}' (tiene encabezado: {has_header}). ---")

        # --- Medición de rendimiento: Inicio total de la función ---
        start_time_total_operation = time.time()
        cell_value: Optional[str] = None # Inicializamos el valor de la celda

        try:
            # Convierte el número de fila lógica (1-basada) a un índice 0-basado para Python
            # Si hay encabezado, la primera fila de datos (lógica 1) está en el índice físico 1 (0-basado).
            # Por lo tanto, si has_header, fila_logica 1 -> índice 1. Sin has_header, fila_logica 1 -> índice 0.
            actual_fila_0_indexed = fila_logica - 1
            if has_header:
                actual_fila_0_indexed += 1 # Ajusta si hay encabezado para saltar la fila 0

            # Convierte el número de columna lógica (1-basada) a un índice 0-basado para Python
            actual_col_0_indexed = columna_logica - 1

            self.logger.info(f"\n🔎 Calculando índices físicos: Fila física (0-indexed): {actual_fila_0_indexed}, Columna física (0-indexed): {actual_col_0_indexed}.")

            # --- Medición de rendimiento: Carga del archivo CSV y lectura de todas las filas ---
            start_time_load_csv = time.time()
            self.logger.info(f"\n⏳ Abriendo y leyendo todas las filas del archivo CSV: '{archivo_csv_path}'...")
            with open(archivo_csv_path, 'r', newline='', encoding='utf-8-sig') as csvfile:
                csv_reader = csv.reader(csvfile, delimiter=delimiter)
                rows = list(csv_reader) # Lee todas las filas del CSV en una lista de listas (cada sublista es una fila)
            end_time_load_csv = time.time()
            duration_load_csv = end_time_load_csv - start_time_load_csv
            self.logger.info(f"PERFORMANCE: Tiempo de carga del archivo CSV y lectura de todas las filas: {duration_load_csv:.4f} segundos.")
            
            self.logger.info(f"\n✅ Archivo CSV leído. Total de filas físicas encontradas: {len(rows)}.")

            # Validación de límites para la fila
            if actual_fila_0_indexed < 0 or actual_fila_0_indexed >= len(rows):
                self.logger.error(f"\n❌ Error: La fila lógica {fila_logica} (física 0-indexed: {actual_fila_0_indexed}) está fuera de los límites del archivo CSV '{archivo_csv_path}'. Total filas físicas: {len(rows)}.")
                return None

            # Validación de límites para la columna en la fila específica
            if actual_col_0_indexed < 0 or actual_col_0_indexed >= len(rows[actual_fila_0_indexed]):
                self.logger.error(f"\n❌ Error: La columna lógica {columna_logica} (física 0-indexed: {actual_col_0_indexed}) está fuera de los límites de la fila física {actual_fila_0_indexed} del archivo CSV '{archivo_csv_path}'. Total columnas en esa fila: {len(rows[actual_fila_0_indexed])}.")
                return None

            # Obtiene el valor de la celda especificada
            cell_value = rows[actual_fila_0_indexed][actual_col_0_indexed]
            
            self.logger.info(f"\n✅ Dato obtenido de (Fila lógica: {fila_logica}, Columna lógica: {columna_logica}) en '{archivo_csv_path}': '{cell_value}'.")
            return cell_value
        
        except FileNotFoundError:
            error_msg = f"\n❌ FALLO (Archivo no encontrado): El archivo CSV no se encontró en la ruta: '{archivo_csv_path}'."
            self.logger.critical(error_msg)
            return None
        except ValueError:
            # Esto ocurriría si fila_logica o columna_logica no fueran enteros,
            # pero los type hints ya lo previenen. Se mantiene por robustez.
            error_msg = f"\n❌ FALLO (Valor inválido): Los parámetros 'fila_logica' y 'columna_logica' deben ser números enteros. Se recibieron: fila='{fila_logica}', columna='{columna_logica}'."
            self.logger.critical(error_msg)
            return None
        except csv.Error as e:
            error_msg = f"\n❌ FALLO (Error de formato CSV): Ocurrió un error al procesar el archivo CSV '{archivo_csv_path}'.\nDetalles: {e}"
            self.logger.critical(error_msg, exc_info=True) # Incluye el stack trace
            return None
        except Exception as e:
            error_msg = (
                f"\n❌ FALLO (Error Inesperado): Ocurrió un error desconocido al leer el dato de la columna del CSV.\n"
                f"Archivo: '{archivo_csv_path}'.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True) # Incluye el stack trace
            return None
        finally:
            # --- Medición de rendimiento: Fin total de la función ---
            end_time_total_operation = time.time()
            duration_total_operation = end_time_total_operation - start_time_total_operation
            self.logger.info(f"PERFORMANCE: Tiempo total de la operación (dato_Columna_csv): {duration_total_operation:.4f} segundos.")
            self.logger.debug("\nFinalizada la operación de lectura de dato de CSV.")
    
    @allure.step("Leyendo archivo JSON: '{json_file_path}'")
    def leer_json(self, json_file_path: str, nombre_paso: str = "") -> Union[Dict, List, None]:
        """
        Lee y parsea un archivo JSON, devolviendo su contenido como un diccionario o lista de Python.
        Esta función mide el tiempo que tarda en abrir, leer y parsear el archivo JSON,
        lo cual es útil para evaluar el rendimiento en escenarios de automatización impulsados por datos.

        Args:
            json_file_path (str): La **ruta completa al archivo JSON**.
            nombre_paso (str, opcional): Una descripción del paso que se está ejecutando para el registro (logs).
                                         Por defecto es una cadena vacía "".

        Returns:
            Union[Dict, List, None]: El contenido del archivo JSON como un **diccionario** o una **lista**,
                                     o **None** si el archivo no se encuentra, el formato JSON es inválido,
                                     o si ocurre un error inesperado.
        """
        nombre_paso = f"Leyendo archivo JSON: '{json_file_path}'"
        self.registrar_paso(nombre_paso)
        
        self.logger.info(f"\n--- {nombre_paso}: Intentando leer el archivo JSON: '{json_file_path}'. ---")

        # --- Medición de rendimiento: Inicio de la operación total de la función ---
        start_time_total_operation = time.time()
        
        data_content: Union[Dict, List, None] = None # Inicializamos a None

        try:
            self.logger.info(f"\n⏳ Abriendo y leyendo el archivo JSON: '{json_file_path}'...")
            with open(json_file_path, 'r', encoding='utf-8-sig') as file:
                # 'encoding='utf-8'' es una buena práctica para manejar caracteres especiales.
                data_content = json.load(file) # Carga (parsea) el contenido del archivo JSON
            
            self.logger.info(f"\n✅ Archivo JSON '{json_file_path}' leído y parseado exitosamente.")
            return data_content

        except FileNotFoundError:
            error_msg = f"\n❌ FALLO (Archivo no encontrado): El archivo JSON no se encontró en la ruta: '{json_file_path}'."
            self.logger.critical(error_msg)
            return None
        except json.JSONDecodeError as e:
            error_msg = f"\n❌ FALLO (Error de formato JSON): Error al decodificar JSON desde '{json_file_path}'.\nDetalles: {e}"
            self.logger.critical(error_msg, exc_info=True) # Incluye el stack trace completo para errores de decodificación JSON
            return None
        except Exception as e:
            error_msg = (
                f"\n❌ FALLO (Error Inesperado): Ocurrió un error inesperado al leer el archivo JSON.\n"
                f"Archivo: '{json_file_path}'.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True) # Incluye el stack trace completo para errores inesperados
            return None
        finally:
            # --- Medición de rendimiento: Fin de la operación total de la función ---
            end_time_total_operation = time.time()
            duration_total_operation = end_time_total_operation - start_time_total_operation
            self.logger.info(f"PERFORMANCE: Tiempo total de la operación (leer_json): {duration_total_operation:.4f} segundos.")
            self.logger.debug("\nOperación de lectura de archivo JSON finalizada.")
            
    @allure.step("Leyendo archivo CSV como diccionario: '{csv_file_path}'")
    def leer_csv_diccionario(self, csv_file_path: str, nombre_paso: str = "") -> Union[List[Dict[str, str]], None]:
        """
        Lee y parsea un archivo CSV, devolviendo su contenido como una lista de diccionarios,
        donde cada diccionario representa una fila y sus claves son los encabezados.
        Esta función mide el tiempo que tarda en abrir, leer y parsear el archivo,
        lo cual es útil para evaluar el rendimiento.

        Args:
            csv_file_path (str): La **ruta completa al archivo CSV**.
            nombre_paso (str, opcional): Una descripción del paso que se está ejecutando para el registro (logs).
                                         Por defecto es una cadena vacía "".

        Returns:
            Union[List[Dict[str, str]], None]: El contenido del archivo CSV como una **lista de diccionarios**,
                                               o **None** si el archivo no se encuentra, el formato CSV es inválido,
                                               o si ocurre un error inesperado.
        """
        nombre_paso = f"Leyendo archivo CSV como diccionario: '{csv_file_path}'"
        self.registrar_paso(nombre_paso)
        
        self.logger.info(f"\n--- {nombre_paso}: Intentando leer el archivo CSV: '{csv_file_path}'. ---")

        # --- Medición de rendimiento: Inicio de la operación total de la función ---
        start_time_total_operation = time.time()
        
        data_content: Union[List[Dict[str, str]], None] = None  # Inicializamos a None

        try:
            self.logger.info(f"\n⏳ Abriendo y leyendo el archivo CSV: '{csv_file_path}'...")
            with open(csv_file_path, mode='r', newline='', encoding='utf-8-sig') as file:
                # csv.DictReader lee cada fila como un diccionario usando los encabezados como claves
                reader = csv.DictReader(file)
                # Convertimos el iterador en una lista de diccionarios
                data_content = list(reader)
            
            self.logger.info(f"\n✅ Archivo CSV '{csv_file_path}' leído y parseado exitosamente.")
            return data_content

        except FileNotFoundError:
            error_msg = f"\n❌ FALLO (Archivo no encontrado): El archivo CSV no se encontró en la ruta: '{csv_file_path}'."
            self.logger.critical(error_msg)
            return None
        except csv.Error as e:
            error_msg = f"\n❌ FALLO (Error de formato CSV): Error al procesar el archivo CSV desde '{csv_file_path}'.\nDetalles: {e}"
            self.logger.critical(error_msg, exc_info=True)  # Incluye el stack trace completo
            return None
        except Exception as e:
            error_msg = (
                f"\n❌ FALLO (Error Inesperado): Ocurrió un error inesperado al leer el archivo CSV.\n"
                f"Archivo: '{csv_file_path}'.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)  # Incluye el stack trace completo
            return None
        finally:
            # --- Medición de rendimiento: Fin de la operación total de la función ---
            end_time_total_operation = time.time()
            duration_total_operation = end_time_total_operation - start_time_total_operation
            self.logger.info(f"PERFORMANCE: Tiempo total de la operación (leer_csv_diccionario): {duration_total_operation:.4f} segundos.")
            self.logger.debug("\nOperación de lectura de archivo CSV finalizada.")
            
    @allure.step("Leyendo archivo Excel/CSV como diccionario: '{excel_file_path}', hoja: '{sheet_name}', con encabezado: {has_header}, con encabezados personalizados: {headers}")
    def leer_excel_diccionario(self, excel_file_path: str, sheet_name: str, has_header: bool = True, headers: Optional[List[str]] = None, nombre_paso: str = "") -> Union[List[Dict[str, Any]], None]:
        """
        Lee y parsea un archivo Excel o CSV, devolviendo su contenido como una lista de diccionarios.
        La función puede manejar archivos con o sin encabezado.

        Args:
            excel_file_path (str): La **ruta completa al archivo**.
            sheet_name (str): El **nombre de la hoja de cálculo** a leer dentro del archivo Excel.
            has_header (bool): Un indicador booleano. Si es True (valor por defecto),
                                la primera fila del archivo se trata como el encabezado.
                                Si es False, los encabezados deben proporcionarse en el
                                parámetro 'headers' y la lectura de datos comenzará desde la primera fila.
            headers (Optional[List[str]]): Una lista de cadenas que representan los encabezados en el
                                            orden correcto. Este parámetro es **obligatorio si has_header es False**.
                                            Por defecto es None.
            nombre_paso (str, opcional): Una descripción del paso que se está ejecutando para el registro (logs).
                                        Por defecto es una cadena vacía "".

        Returns:
            Union[List[Dict[str, Any]], None]: El contenido del archivo como una **lista de diccionarios**,
                                            o **None** si el archivo no se encuentra, el formato es inválido,
                                            si la hoja no existe o si ocurre un error inesperado.
        """
        nombre_paso = f"Leyendo archivo Excel/CSV como diccionario: '{excel_file_path}', hoja: '{sheet_name}', con encabezado: {has_header}, con encabezados personalizados: {headers}"
        self.registrar_paso(nombre_paso)
        
        self.logger.info(f"\n--- {nombre_paso}: Intentando leer el archivo: '{excel_file_path}'. ---")
        start_time_total_operation = time.time()
        data_content: List[Dict[str, Any]] = []

        try:
            # Manejo para archivos CSV
            if excel_file_path.endswith('.csv'):
                self.logger.info(f"\n⏳ Abriendo y leyendo el archivo CSV: '{excel_file_path}'...")
                with open(excel_file_path, mode='r', newline='', encoding='utf-8-sig') as file:
                    reader = csv.reader(file)
                    start_row = 1
                    
                    if has_header:
                        headers = next(reader)
                    elif not has_header and not headers:
                        raise ValueError("El parámetro 'headers' es obligatorio si 'has_header' es False.")
                    
                    for row in reader:
                        row_dict = dict(zip(headers, row))
                        data_content.append(row_dict)

            # Manejo para archivos Excel (.xlsx)
            else:
                self.logger.info(f"\n⏳ Abriendo y leyendo el archivo Excel: '{excel_file_path}'...")
                workbook = openpyxl.load_workbook(excel_file_path)
                
                # Se utiliza el nombre de la hoja proporcionado por el usuario
                sheet = workbook[sheet_name]
                
                start_row = 1
                if has_header:
                    headers = [cell.value for cell in sheet[1]]
                    start_row = 2
                elif not has_header and not headers:
                    raise ValueError("El parámetro 'headers' es obligatorio si 'has_header' es False.")
                
                for row_idx, row in enumerate(sheet.iter_rows(min_row=start_row)):
                    row_dict = {}
                    for header, cell in zip(headers, row):
                        if header is not None:
                            row_dict[header] = cell.value
                    data_content.append(row_dict)

            self.logger.info(f"\n✅ Archivo '{excel_file_path}' leído y parseado exitosamente.")
            return data_content

        except FileNotFoundError:
            error_msg = f"\n❌ FALLO (Archivo no encontrado): El archivo no se encontró en la ruta: '{excel_file_path}'."
            self.logger.critical(error_msg)
            return None
        except KeyError:
            error_msg = f"\n❌ FALLO (Hoja no encontrada): La hoja de cálculo '{sheet_name}' no se encontró en el archivo: '{excel_file_path}'."
            self.logger.critical(error_msg)
            return None
        except (BadZipFile, InvalidFileException):
            error_msg = f"\n❌ FALLO (Error de formato de archivo): El archivo '{excel_file_path}' no es un archivo válido."
            self.logger.critical(error_msg, exc_info=True)
            return None
        except Exception as e:
            error_msg = (
                f"\n❌ FALLO (Error Inesperado): Ocurrió un error inesperado al leer el archivo.\n"
                f"Archivo: '{excel_file_path}'.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            return None
        finally:
            end_time_total_operation = time.time()
            duration_total_operation = end_time_total_operation - start_time_total_operation
            self.logger.info(f"PERFORMANCE: Tiempo total de la operación (leer_excel_diccionario): {duration_total_operation:.4f} segundos.")
            self.logger.debug("\nOperación de lectura de archivo finalizada.")
        
    @allure.step("Leyendo archivo de texto plano: '{file_path}', con delimitador: '{delimiter}'")
    def leer_texto_plano(self, file_path: str, delimiter: Optional[str] = None, nombre_paso: str = "") -> Union[str, List[str], None]:
        """
        Lee el contenido completo de un archivo de texto plano.
        Si se proporciona un delimitador, divide el contenido del archivo por el delimitador
        y lo devuelve como una lista de cadenas.
        Esta función mide el tiempo que tarda en abrir, leer y procesar el archivo de texto,
        lo cual es útil para evaluar el rendimiento en operaciones de E/S de archivos.

        Args:
            file_path (str): La **ruta completa al archivo de texto**.
            delimiter (str, opcional): Si se proporciona, el contenido del archivo se dividirá por este delimitador
                                        y se devolverá como una lista de cadenas. Si es `None`, se devuelve el contenido
                                        completo como una sola cadena. Por defecto es `None`.
            nombre_paso (str, opcional): Una descripción del paso que se está ejecutando para el registro (logs).
                                         Por defecto es una cadena vacía "".

        Returns:
            Union[str, List[str], None]: El contenido del archivo como una **cadena** (si `delimiter` es `None`)
                                         o una **lista de cadenas** (si se usa `delimiter`).
                                         Retorna `None` si el archivo no se encuentra, hay un error de E/S,
                                         o si ocurre un error inesperado.
        """
        delimiter_log_info = f"'{delimiter}'" if delimiter is not None else "Ninguno"
        
        nombre_paso = f"Leyendo archivo de texto plano: '{file_path}', con delimitador: '{delimiter_log_info}'"
        self.registrar_paso(nombre_paso)
        
        self.logger.info(f"\n--- {nombre_paso}: Intentando leer el archivo de texto: '{file_path}' (Delimitador: {delimiter_log_info}). ---")

        # --- Medición de rendimiento: Inicio de la operación total de la función ---
        start_time_total_operation = time.time()
        
        content: Optional[str] = None # Inicializamos content

        try:
            self.logger.info(f"\n⏳ Abriendo y leyendo el archivo de texto: '{file_path}'...")
            with open(file_path, 'r', encoding='utf-8-sig') as file:
                # 'encoding='utf-8'' es crucial para manejar correctamente una amplia gama de caracteres.
                content = file.read() # Lee todo el contenido del archivo
            
            self.logger.info(f"\n✅ Archivo de texto '{file_path}' leído exitosamente.")

            if delimiter is not None:
                # --- Medición de rendimiento: División del contenido (si aplica) ---
                start_time_split = time.time()
                self.logger.info(f"\n🔎 Dividiendo el contenido por el delimitador: '{delimiter}'...")
                result = content.split(delimiter) # Divide el contenido por el delimitador y lo retorna como lista
                end_time_split = time.time()
                duration_split = end_time_split - start_time_split
                self.logger.info(f"PERFORMANCE: Tiempo de división del contenido: {duration_split:.4f} segundos.")
                self.logger.info(f"\n✅ Archivo de texto '{file_path}' leído y dividido exitosamente. Se encontraron {len(result)} segmentos.")
                return result
            else:
                self.logger.info(f"\n✅ Archivo de texto '{file_path}' leído completamente como una sola cadena.")
                return content
            
        except FileNotFoundError:
            error_msg = f"\n❌ FALLO (Archivo no encontrado): El archivo de texto no se encontró en la ruta: '{file_path}'."
            self.logger.critical(error_msg)
            return None
        except IOError as e:
            error_msg = f"\n❌ FALLO (Error de E/S): Ocurrió un error de entrada/salida al leer el archivo de texto '{file_path}'.\nDetalles: {e}"
            self.logger.critical(error_msg, exc_info=True) # Incluye el stack trace completo
            return None
        except Exception as e:
            error_msg = (
                f"\n❌ FALLO (Error Inesperado): Ocurrió un error desconocido al leer el archivo de texto.\n"
                f"Archivo: '{file_path}'.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True) # Incluye el stack trace completo
            return None
        finally:
            # --- Medición de rendimiento: Fin de la operación total de la función ---
            end_time_total_operation = time.time()
            duration_total_operation = end_time_total_operation - start_time_total_operation
            self.logger.info(f"PERFORMANCE: Tiempo total de la operación (leer_texto): {duration_total_operation:.4f} segundos.")
            self.logger.debug("\nOperación de lectura de archivo de texto finalizada.")

    @allure.step("Leyendo archivo XML: '{xml_file_path}'")
    def leer_xml(self, xml_file_path: str, nombre_paso: str = "") -> Union[ET.Element, None]:
        """
        Lee y parsea un archivo XML, devolviendo su elemento raíz como un objeto Element.
        Esta función mide el tiempo que tarda en abrir, leer y parsear el archivo XML,
        lo cual es útil para evaluar el rendimiento en escenarios donde se procesan archivos XML.

        Args:
            xml_file_path (str): La **ruta completa al archivo XML**.
            nombre_paso (str, opcional): Una descripción del paso que se está ejecutando para el registro (logs).
                                         Por defecto es una cadena vacía "".

        Returns:
            Union[ET.Element, None]: El **elemento raíz del XML** como un objeto `xml.etree.ElementTree.Element`,
                                     o **None** si el archivo no se encuentra, el formato XML es inválido,
                                     o si ocurre un error inesperado.
        """
        nombre_paso = f"Leyendo archivo XML: '{xml_file_path}'"
        self.registrar_paso(nombre_paso)
        
        self.logger.info(f"\n--- {nombre_paso}: Intentando leer el archivo XML: '{xml_file_path}'. ---")

        # --- Medición de rendimiento: Inicio de la operación total de la función ---
        start_time_total_operation = time.time()
        
        root_element: Optional[ET.Element] = None # Inicializamos el elemento raíz

        try:
            self.logger.info(f"\n⏳ Abriendo y parseando el archivo XML: '{xml_file_path}'...")
            # ET.parse() se encarga de abrir y parsear el archivo.
            # No es necesario especificar la codificación en la mayoría de los casos ya que
            # ET lo detecta automáticamente si el XML tiene una declaración de codificación (e.g., <?xml version="1.0" encoding="UTF-8"?>).
            tree = ET.parse(xml_file_path)
            
            # Obtiene el elemento raíz del XML
            root_element = tree.getroot()
            
            self.logger.info(f"\n✅ Archivo XML '{xml_file_path}' leído y parseado exitosamente. Elemento raíz: '{root_element.tag}'.")
            return root_element

        except FileNotFoundError:
            error_msg = f"\n❌ FALLO (Archivo no encontrado): El archivo XML no se encontró en la ruta: '{xml_file_path}'."
            self.logger.critical(error_msg)
            return None
        except ET.ParseError as e:
            error_msg = f"\n❌ FALLO (Error de formato XML): Ocurrió un error al parsear el archivo XML '{xml_file_path}'.\nDetalles: {e}"
            self.logger.critical(error_msg, exc_info=True) # Incluye el stack trace completo para errores de parseo XML
            return None
        except Exception as e:
            error_msg = (
                f"\n❌ FALLO (Error Inesperado): Ocurrió un error desconocido al leer el archivo XML.\n"
                f"Archivo: '{xml_file_path}'.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True) # Incluye el stack trace completo para errores inesperados
            return None
        finally:
            # --- Medición de rendimiento: Fin de la operación total de la función ---
            end_time_total_operation = time.time()
            duration_total_operation = end_time_total_operation - start_time_total_operation
            self.logger.info(f"PERFORMANCE: Tiempo total de la operación (leer_xml): {duration_total_operation:.4f} segundos.")
            self.logger.debug("\nOperación de lectura de archivo XML finalizada.")
    
    @allure.step("Escribiendo en archivo de texto plano: '{file_path}', append: {append}, delimitador: '{delimiter}'")
    def escribir_texto_plano(self, file_path: str, content: Union[str, List[str]], append: bool = False, delimiter: Optional[str] = None, nombre_paso: str = "") -> bool:
        """
        Escribe contenido en un archivo de texto plano. Si el contenido es una lista de cadenas
        y se proporciona un delimitador, las cadenas se unirán con el delimitador antes de escribirlas.
        Esta función mide el tiempo de preparación del contenido y la escritura en el archivo,
        lo cual es útil para evaluar el rendimiento de las operaciones de E/S.

        Args:
            file_path (str): La **ruta completa al archivo de texto**.
            content (Union[str, List[str]]): La cadena o lista de cadenas a escribir.
            append (bool, opcional): Si es `True`, el contenido se añadirá al final del archivo.
                                     Si es `False` (por defecto), el archivo se sobrescribirá si existe.
            delimiter (str, opcional): Si se proporciona y `content` es una lista de cadenas, las cadenas
                                       se unirán con este delimitador antes de la escritura. Si es `None`,
                                       las cadenas de una lista se escribirán directamente sin separación explícita.
                                       Por defecto es `None`.
            nombre_paso (str, opcional): Una descripción del paso que se está ejecutando para el registro (logs).
                                         Por defecto es una cadena vacía "".

        Returns:
            bool: `True` si la escritura fue exitosa, `False` en caso de error.
        """
        mode = 'a' if append else 'w' # Determina el modo de apertura: 'a' para añadir, 'w' para sobrescribir
        action = "añadir a" if append else "escribir en" # Descripción de la acción para el log
        
        delimiter_log_info = f"'{delimiter}'" if delimiter is not None else "Ninguno"
        
        nombre_paso = f"Escribiendo en archivo de texto plano: '{file_path}', append: {action}, delimitador: '{delimiter_log_info}'"
        self.registrar_paso(nombre_paso)
        
        self.logger.info(f"\n--- {nombre_paso}: Intentando {action} el archivo de texto: '{file_path}' (Delimitador de escritura: {delimiter_log_info}). ---")

        # --- Medición de rendimiento: Inicio de la operación total de la función ---
        start_time_total_operation = time.time()
        
        text_to_write: str = "" # Variable para almacenar el contenido final a escribir

        try:
            # Lógica para procesar el contenido antes de la escritura
            if isinstance(content, list):
                # --- Medición de rendimiento: Unión de la lista con el delimitador ---
                start_time_join = time.time()
                
                if delimiter is not None:
                    text_to_write = delimiter.join(content)
                    self.logger.info(f"\n🔎 El contenido de la lista será unido con el delimitador '{delimiter}' antes de escribir.")
                else:
                    text_to_write = "".join(content)
                    self.logger.warning("\n⚠️ Se proporcionó una lista para escribir_texto sin delimitador. Las cadenas se concatenarán sin separación explícita, lo que puede no ser el comportamiento deseado.")
                
                end_time_join = time.time()
                duration_join = end_time_join - start_time_join
                self.logger.info(f"PERFORMANCE: Tiempo de preparación del contenido (join): {duration_join:.4f} segundos.")

            elif isinstance(content, str):
                text_to_write = content # Si el contenido ya es una cadena, lo asigna tal cual
                self.logger.info("\n🔎 El contenido es una cadena, se escribirá directamente.")
            else:
                error_msg = f"\n❌ FALLO (Tipo de dato inválido): El tipo de contenido proporcionado no es válido. Se esperaba str o List[str], se recibió: {type(content)}."
                self.logger.critical(error_msg)
                return False

            # --- Medición de rendimiento: Escritura en el archivo ---
            self.logger.info(f"\n✍️ Escribiendo contenido en el archivo: '{file_path}'...")
            with open(file_path, mode, encoding='utf-8-sig') as file:
                # `encoding='utf-8'` es crucial para manejar correctamente una amplia gama de caracteres
                file.write(text_to_write)
            
            self.logger.info(f"\n✅ Contenido {action} exitosamente en '{file_path}'.")
            return True
        
        except IOError as e:
            error_msg = f"\n❌ FALLO (Error de E/S): Ocurrió un error de entrada/salida al {action} el archivo de texto '{file_path}'.\nDetalles: {e}"
            self.logger.critical(error_msg, exc_info=True) # Incluye el stack trace completo
            return False
        except Exception as e:
            error_msg = (
                f"\n❌ FALLO (Error Inesperado): Ocurrió un error desconocido al {action} el archivo de texto.\n"
                f"Archivo: '{file_path}'.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True) # Incluye el stack trace completo
            return False
        finally:
            # --- Medición de rendimiento: Fin de la operación total de la función ---
            end_time_total_operation = time.time()
            duration_total_operation = end_time_total_operation - start_time_total_operation
            self.logger.info(f"PERFORMANCE: Tiempo total de la operación (escribir_texto): {duration_total_operation:.4f} segundos.")
            self.logger.debug("\nOperación de escritura de archivo de texto finalizada.")
    
    @allure.step("Escribiendo en archivo JSON: '{file_path}', append: {append}, indent: {indent}")
    def escribir_json(self, file_path: str, data: Union[Dict, List], indent: int = 4, append: bool = False, nombre_paso: str = "") -> bool:
        """
        Escribe un objeto Python (diccionario o lista) en un archivo JSON.
        Esta función serializa el objeto en una cadena JSON y lo guarda en el disco,
        con formato legible por defecto.

        Args:
            file_path (str): La ruta completa al archivo JSON.
            data (Union[Dict, List]): El diccionario o lista de datos a escribir en el archivo.
            indent (int, opcional): El número de espacios de sangría para dar formato al JSON.
                                    Por defecto es 4.
            append (bool, opcional): Si es `True`, los datos se añadirán a una lista existente en el archivo.
                                    Si es `False` (por defecto), el archivo se sobrescribirá.
            nombre_paso (str, opcional): Una descripción del paso que se está ejecutando para el registro (logs).
                                        Por defecto es una cadena vacía "".

        Returns:
            bool: `True` si la escritura fue exitosa, `False` en caso de error.
        """
        mode_action = "añadir a" if append else "escribir en"
        
        nombre_paso = f"Escribiendo en archivo JSON: '{file_path}', append: {mode_action}, indent: {indent}"
        self.registrar_paso(nombre_paso)
        
        self.logger.info(f"\n--- {nombre_paso}: Intentando {mode_action} el archivo JSON: '{file_path}'. ---")
        
        # --- Medición de rendimiento: Inicio de la operación total de la función ---
        start_time_total_operation = time.time()
        
        try:
            final_data = data
            if append:
                # Comprueba si el archivo existe y tiene contenido
                if os.path.exists(file_path) and os.path.getsize(file_path) > 0:
                    with open(file_path, 'r', encoding='utf-8-sig') as file:
                        existing_data = json.load(file)
                    
                    if isinstance(existing_data, list):
                        if isinstance(data, list):
                            final_data = existing_data + data
                        else:
                            final_data = existing_data + [data]
                        self.logger.info(f"\n🔎 Modo 'append': Se añadieron nuevos datos a la lista existente del archivo.")
                    else:
                        raise TypeError(f"El modo 'append' requiere que el archivo JSON contenga una lista, pero se encontró un tipo '{type(existing_data).__name__}'.")
                else:
                    # Si el archivo no existe o está vacío, crea una nueva lista
                    if not isinstance(data, list):
                        final_data = [data]
                    self.logger.info(f"\n🔎 El archivo no existe o está vacío. Se creó un nuevo archivo con los datos iniciales.")
            
            # --- Medición de rendimiento: Serialización a JSON ---
            start_time_serialization = time.time()
            json_string = json.dumps(final_data, indent=indent, ensure_ascii=False)
            
            end_time_serialization = time.time()
            duration_serialization = end_time_serialization - start_time_serialization
            self.logger.info(f"PERFORMANCE: Tiempo de serialización del objeto a JSON: {duration_serialization:.4f} segundos.")

            # --- Medición de rendimiento: Escritura en el archivo ---
            self.logger.info(f"\n✍️ Escribiendo contenido JSON en el archivo: '{file_path}'...")
            with open(file_path, 'w', encoding='utf-8-sig') as file:
                file.write(json_string)
            
            self.logger.info(f"\n✅ Contenido JSON {mode_action} exitosamente en '{file_path}'.")
            return True
        
        except (TypeError, json.JSONDecodeError) as e:
            error_msg = f"\n❌ FALLO (Error de Formato): No se pudo serializar o decodificar el JSON. Asegúrate de que el archivo JSON esté bien formado y que los datos sean compatibles.\nDetalles: {e}"
            self.logger.critical(error_msg, exc_info=True)
            return False
            
        except IOError as e:
            error_msg = f"\n❌ FALLO (Error de E/S): Ocurrió un error de entrada/salida al escribir el archivo '{file_path}'.\nDetalles: {e}"
            self.logger.critical(error_msg, exc_info=True)
            return False
            
        except Exception as e:
            error_msg = f"\n❌ FALLO (Error Inesperado): Ocurrió un error desconocido al escribir el archivo JSON.\nArchivo: '{file_path}'.\nDetalles: {e}"
            self.logger.critical(error_msg, exc_info=True)
            return False
            
        finally:
            # --- Medición de rendimiento: Fin de la operación total de la función ---
            end_time_total_operation = time.time()
            duration_total_operation = end_time_total_operation - start_time_total_operation
            self.logger.info(f"PERFORMANCE: Tiempo total de la operación (escribir_json): {duration_total_operation:.4f} segundos.")
            self.logger.debug("\nOperación de escritura de archivo JSON finalizada.")
    
    @allure.step("Escribiendo en archivo Excel: '{file_path}', append: {append}, header: {header}")
    def escribir_excel(self, file_path: str, data: List[Dict], append: bool = False, header: bool = True, nombre_paso: str = "") -> bool:
        """
        Escribe una lista de diccionarios en un archivo Excel.
        Permite sobrescribir el archivo o añadir datos a uno existente.
        También valida si la cabecera existe y coincide al añadir datos.

        Args:
            file_path (str): La ruta completa al archivo Excel (.xlsx).
            data (List[Dict]): Una lista de diccionarios con los datos a escribir.
                               Cada diccionario representa una fila y las claves, los encabezados.
            append (bool, opcional): Si es `True`, los datos se añadirán a un archivo existente.
                                     Si es `False` (por defecto), el archivo se sobrescribirá.
            header (bool, opcional): Si es `True`, se incluirá la cabecera de las columnas.
                                     Si es `False`, solo se escribirán los datos. Por defecto es `True`.
            nombre_paso (str, opcional): Una descripción del paso que se está ejecutando para el registro (logs).
                                         Por defecto es una cadena vacía "".

        Returns:
            bool: `True` si la escritura fue exitosa, `False` en caso de error.
        """
        mode_action = "añadir a" if append else "escribir en"
        
        nombre_paso = f"Escribiendo en archivo Excel: '{file_path}', append: {mode_action}, header: {header}"
        self.registrar_paso(nombre_paso)
        
        self.logger.info(f"\n--- {nombre_paso}: Intentando {mode_action} el archivo Excel: '{file_path}'. ---")
        start_time_total_operation = time.time()

        try:
            if not isinstance(data, list) or not all(isinstance(d, dict) for d in data):
                raise TypeError("Los datos para escribir en Excel deben ser una lista de diccionarios.")

            df_new = pd.DataFrame(data)

            if append and os.path.exists(file_path) and os.path.getsize(file_path) > 0:
                self.logger.info(f"\n🔎 Modo 'append': Verificando archivo existente para añadir datos.")
                
                df_existing = pd.read_excel(file_path)
                
                # Validación de cabeceras solo si se especifica que se quiere una cabecera
                if header:
                    existing_headers = df_existing.columns.tolist()
                    new_headers = df_new.columns.tolist()
                    if existing_headers != new_headers:
                        error_msg = f"\n❌ ERROR: Las cabeceras del archivo existente no coinciden con las nuevas. No se puede añadir la data. \n Cabeceras existentes: {existing_headers} \n Cabeceras nuevas: {new_headers}"
                        self.logger.critical(error_msg)
                        return False
                
                start_row = len(df_existing) + 1
                
                with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                    df_new.to_excel(writer, index=False, header=False, sheet_name='Sheet1', startrow=start_row)
            else:
                self.logger.info(f"\n🔎 El archivo no existe o se sobrescribirá. Escribiendo nuevos datos.")
                df_new.to_excel(file_path, index=False, header=header, sheet_name='Sheet1')
            
            self.logger.info(f"\n✅ Contenido Excel {mode_action} exitosamente en '{file_path}'.")
            return True

        except pd.errors.EmptyDataError:
            error_msg = f"\n❌ FALLO (Error de Datos): El archivo '{file_path}' está vacío o no tiene el formato correcto."
            self.logger.critical(error_msg)
            return False
        except IOError as e:
            error_msg = f"\n❌ FALLO (Error de E/S): Ocurrió un error de entrada/salida al {mode_action} el archivo Excel '{file_path}'.\nDetalles: {e}"
            self.logger.critical(error_msg, exc_info=True)
            return False
        except Exception as e:
            error_msg = f"\n❌ FALLO (Error Inesperado): Ocurrió un error desconocido al escribir el archivo Excel.\nArchivo: '{file_path}'.\nDetalles: {e}"
            self.logger.critical(error_msg, exc_info=True)
            return False
        finally:
            end_time_total_operation = time.time()
            duration_total_operation = end_time_total_operation - start_time_total_operation
            self.logger.info(f"PERFORMANCE: Tiempo total de la operación (escribir_excel): {duration_total_operation:.4f} segundos.")
            self.logger.debug("\nOperación de escritura de archivo Excel finalizada.")
            
    @allure.step("Escribiendo en archivo CSV: '{file_path}', append: {append}, header: {header}")
    def escribir_csv(self, file_path: str, data: List[Dict], append: bool = False, header: bool = True, nombre_paso: str = "escribir_csv") -> bool:
        """
        Escribe datos en un archivo CSV.

        Esta función ahora es más robusta. Si se le pide anexar datos a un archivo existente,
        primero leerá los encabezados del archivo para asegurar que la nueva fila se
        escribe en el orden correcto, evitando el problema de desalineación.

        Args:
            file_path (str): Ruta completa del archivo CSV a escribir.
            data (List[Dict]): Una lista de diccionarios, donde cada diccionario representa una fila.
            append (bool): Si es True, anexa los datos al final del archivo.
            header (bool): Si es True, escribe los encabezados de las columnas.
            nombre_paso (str): Nombre descriptivo para el paso de log.

        Returns:
            bool: True si la escritura fue exitosa, False en caso de error.
        """
        mode_desc = 'anexando' if append else 'sobrescribiendo'
        
        nombre_paso = f"Escribiendo en archivo CSV: '{file_path}', append: {mode_desc}, header: {header}"
        self.registrar_paso(nombre_paso)
        
        try:
            # Crea el directorio si no existe.
            dir_name = os.path.dirname(file_path)
            if dir_name and not os.path.exists(dir_name):
                os.makedirs(dir_name)

            fieldnames = data[0].keys() if data else []
            mode = 'a' if append and os.path.exists(file_path) else 'w'
            write_header = not (append and os.path.exists(file_path))

            with open(file_path, mode, newline='', encoding='utf-8-sig') as file:
                writer = csv.DictWriter(file, fieldnames=fieldnames)

                if write_header:
                    writer.writeheader()

                writer.writerows(data)

            self.logger.info(f"✅ {nombre_paso}: Datos escritos correctamente en el archivo CSV: '{file_path}'.")
            return True

        except IOError as e:
            self.logger.error(f"❌ {nombre_paso}: Error de E/S al escribir en el archivo CSV: {e}")
            return False
        except Exception as e:
            self.logger.error(f"❌ {nombre_paso}: Ocurrió un error inesperado: {e}", exc_info=True)
            return False
        
    @allure.step("Modificando registro en archivo: '{file_path}'")
    def modificar_registro(self, file_path: str, clave_busqueda: str, valor_busqueda: Any, nuevos_datos: Dict[str, Any]) -> bool:
        """
        Modifica uno o varios campos (dato específico o datos completos) de un registro
        en un archivo (CSV, JSON, XLSX) basado en una clave y valor.

        Args:
            file_path (str): Ruta completa al archivo.
            clave_busqueda (str): Nombre de la columna/clave para buscar el registro.
            valor_busqueda (Any): Valor que debe coincidir con la clave_busqueda.
            nuevos_datos (Dict[str, Any]): Diccionario con los pares clave-valor a modificar.

        Returns:
            bool: True si el registro fue encontrado y modificado, False en caso contrario o error.
        """
        nombre_paso = f"Modificando registro en '{file_path}' donde {clave_busqueda}='{valor_busqueda}'"
        self.registrar_paso(nombre_paso)
        
        file_extension = os.path.splitext(file_path)[1].lower()
        
        try:
            if file_extension == '.csv':
                return self._modificar_registro_csv(file_path, clave_busqueda, valor_busqueda, nuevos_datos, nombre_paso)
            elif file_extension == '.json':
                return self._modificar_registro_json(file_path, clave_busqueda, valor_busqueda, nuevos_datos, nombre_paso)
            elif file_extension in ['.xlsx', '.xls']:
                return self._modificar_registro_excel(file_path, clave_busqueda, valor_busqueda, nuevos_datos, nombre_paso)
            else:
                self.logger.error(f"\n❌ {nombre_paso}: Tipo de archivo '{file_extension}' no soportado para modificación de registros.")
                return False
        except Exception as e:
            self.logger.error(f"\n❌ {nombre_paso}: Error general durante la modificación. Detalles: {e}", exc_info=True)
            return False

    def _modificar_registro_csv(self, file_path: str, clave_busqueda: str, valor_busqueda: Any, nuevos_datos: Dict[str, Any], nombre_paso: str) -> bool:
        try:
            with open(file_path, mode='r', newline='', encoding='utf-8-sig') as file:
                reader = csv.DictReader(file)
                data = list(reader)
                fieldnames = reader.fieldnames
        except FileNotFoundError:
            self.logger.error(f"\n❌ {nombre_paso}: Archivo CSV no encontrado.")
            return False

        registro_modificado = False
        for row in data:
            if row.get(clave_busqueda) == str(valor_busqueda):
                for key, value in nuevos_datos.items():
                    if key in row:
                        row[key] = str(value)
                registro_modificado = True
                self.logger.info(f"\n✨ Registro encontrado y modificado.")

        if not registro_modificado:
            self.logger.warning(f"\n⚠️ {nombre_paso}: No se encontró registro para modificar.")
            return False

        try:
            with open(file_path, mode='w', newline='', encoding='utf-8-sig') as file:
                writer = csv.DictWriter(file, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(data)
            self.logger.info(f"✅ {nombre_paso}: Archivo CSV reescrito correctamente.")
            return True
        except IOError as e:
            self.logger.error(f"\n❌ {nombre_paso}: Error al reescribir CSV. Detalles: {e}", exc_info=True)
            return False

    def _modificar_registro_json(self, file_path: str, clave_busqueda: str, valor_busqueda: Any, nuevos_datos: Dict[str, Any], nombre_paso: str) -> bool:
        try:
            with open(file_path, 'r', encoding='utf-8-sig') as file:
                data = json.load(file)
        except (FileNotFoundError, json.JSONDecodeError) as e:
            self.logger.error(f"\n❌ {nombre_paso}: Error al leer o decodificar JSON. Detalles: {e}")
            return False
        
        if not isinstance(data, list):
            self.logger.error(f"\n❌ {nombre_paso}: El archivo JSON no es una lista de registros. No se puede modificar.")
            return False

        registro_modificado = False
        for row in data:
            # Comparamos directamente, JSON respeta tipos (a diferencia de CSV)
            if row.get(clave_busqueda) == valor_busqueda:
                for key, value in nuevos_datos.items():
                    if key in row:
                        row[key] = value
                registro_modificado = True
                self.logger.info(f"\n✨ Registro encontrado y modificado.")
                
        if not registro_modificado:
            self.logger.warning(f"\n⚠️ {nombre_paso}: No se encontró registro para modificar.")
            return False

        try:
            with open(file_path, 'w', encoding='utf-8-sig') as file:
                json.dump(data, file, indent=4)
            self.logger.info(f"\n✅ {nombre_paso}: Archivo JSON reescrito correctamente.")
            return True
        except IOError as e:
            self.logger.error(f"\n❌ {nombre_paso}: Error al reescribir JSON. Detalles: {e}", exc_info=True)
            return False
    
    def _modificar_registro_excel(self, file_path: str, clave_busqueda: str, valor_busqueda: Any, nuevos_datos: Dict[str, Any], nombre_paso: str) -> bool:
        try:
            # Usamos Pandas para la manipulación sencilla de Excel
            df = pd.read_excel(file_path)
        except (FileNotFoundError, Exception) as e:
            self.logger.error(f"\n❌ {nombre_paso}: Error al leer archivo Excel con pandas: {e}", exc_info=True)
            return False
        
        # Identifica las filas a modificar (convertimos a str para comparación robusta)
        filas_a_modificar = df[df[clave_busqueda].astype(str) == str(valor_busqueda)].index
        
        if filas_a_modificar.empty:
            self.logger.warning(f"\n⚠️ {nombre_paso}: No se encontró registro para modificar en Excel.")
            return False
        
        # Aplica las modificaciones
        for key, value in nuevos_datos.items():
            if key in df.columns:
                df.loc[filas_a_modificar, key] = value
                self.logger.info(f"\n✨ Columna '{key}' modificada en {len(filas_a_modificar)} registro(s).")
            else:
                self.logger.warning(f"\n⚠️ Columna '{key}' a modificar no existe en el archivo Excel.")
        
        # Reescribe el archivo Excel
        try:
            df.to_excel(file_path, index=False) # index=False para no escribir el índice de pandas
            self.logger.info(f"\n✅ {nombre_paso}: Archivo Excel reescrito correctamente con las modificaciones.")
            return True
        except Exception as e:
            self.logger.error(f"\n❌ {nombre_paso}: Error al reescribir archivo Excel. Detalles: {e}", exc_info=True)
            return False

    @allure.step("Eliminando registro en archivo: '{file_path}'")
    def eliminar_registro(self, file_path: str, clave_busqueda: str, valor_busqueda: Any) -> bool:
        """
        Elimina un registro específico en un archivo (CSV, JSON, XLSX) basado en una clave y valor.

        Args:
            file_path (str): Ruta completa al archivo.
            clave_busqueda (str): Nombre de la columna/clave para buscar el registro a eliminar.
            valor_busqueda (Any): Valor que debe coincidir con la clave_busqueda.

        Returns:
            bool: True si el registro fue eliminado, False en caso contrario o error.
        """
        nombre_paso = f"Eliminando registro en '{file_path}' donde {clave_busqueda}='{valor_busqueda}'"
        self.registrar_paso(nombre_paso)
        
        file_extension = os.path.splitext(file_path)[1].lower()
        
        try:
            if file_extension == '.csv':
                return self._eliminar_registro_csv(file_path, clave_busqueda, valor_busqueda, nombre_paso)
            elif file_extension == '.json':
                return self._eliminar_registro_json(file_path, clave_busqueda, valor_busqueda, nombre_paso)
            elif file_extension in ['.xlsx', '.xls']:
                return self._eliminar_registro_excel(file_path, clave_busqueda, valor_busqueda, nombre_paso)
            else:
                self.logger.error(f"\n❌ {nombre_paso}: Tipo de archivo '{file_extension}' no soportado para eliminación de registros.")
                return False
        except Exception as e:
            self.logger.error(f"\n❌ {nombre_paso}: Error general durante la eliminación. Detalles: {e}", exc_info=True)
            return False

    def _eliminar_registro_csv(self, file_path: str, clave_busqueda: str, valor_busqueda: Any, nombre_paso: str) -> bool:
        try:
            with open(file_path, mode='r', newline='', encoding='utf-8-sig') as file:
                reader = csv.DictReader(file)
                data_original = list(reader)
                fieldnames = reader.fieldnames
        except FileNotFoundError:
            self.logger.error(f"\n❌ {nombre_paso}: Archivo CSV no encontrado.")
            return False
        
        nueva_data = [
            row for row in data_original if row.get(clave_busqueda) != str(valor_busqueda)
        ]

        if len(nueva_data) == len(data_original):
            self.logger.warning(f"\n⚠️ {nombre_paso}: No se encontró registro para eliminar.")
            return False
        
        try:
            with open(file_path, mode='w', newline='', encoding='utf-8-sig') as file:
                writer = csv.DictWriter(file, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(nueva_data)
            self.logger.info(f"\n✅ {nombre_paso}: Registro(s) eliminado(s) y archivo CSV reescrito correctamente.")
            return True
        except IOError as e:
            self.logger.error(f"\n❌ {nombre_paso}: Error al reescribir CSV. Detalles: {e}", exc_info=True)
            return False

    def _eliminar_registro_json(self, file_path: str, clave_busqueda: str, valor_busqueda: Any, nombre_paso: str) -> bool:
        try:
            with open(file_path, 'r', encoding='utf-8-sig') as file:
                data_original = json.load(file)
        except (FileNotFoundError, json.JSONDecodeError) as e:
            self.logger.error(f"\n❌ {nombre_paso}: Error al leer o decodificar JSON. Detalles: {e}")
            return False
        
        if not isinstance(data_original, list):
            self.logger.error(f"\n❌ {nombre_paso}: El archivo JSON no es una lista de registros. No se puede eliminar.")
            return False

        nueva_data = [
            row for row in data_original if row.get(clave_busqueda) != valor_busqueda
        ]
        
        if len(nueva_data) == len(data_original):
            self.logger.warning(f"\n⚠️ {nombre_paso}: No se encontró registro para eliminar.")
            return False

        try:
            with open(file_path, 'w', encoding='utf-8-sig') as file:
                json.dump(nueva_data, file, indent=4)
            self.logger.info(f"\n✅ {nombre_paso}: Registro(s) eliminado(s) y archivo JSON reescrito correctamente.")
            return True
        except IOError as e:
            self.logger.error(f"\n❌ {nombre_paso}: Error al reescribir JSON. Detalles: {e}", exc_info=True)
            return False
            
    def _eliminar_registro_excel(self, file_path: str, clave_busqueda: str, valor_busqueda: Any, nombre_paso: str) -> bool:
        try:
            df = pd.read_excel(file_path)
        except (FileNotFoundError, Exception) as e:
            self.logger.error(f"\n❌ {nombre_paso}: Error al leer archivo Excel con pandas: {e}", exc_info=True)
            return False
        
        df_original_len = len(df)
        
        # Filtra la data: mantiene solo las filas donde el valor NO coincide.
        df_filtrado = df[df[clave_busqueda].astype(str) != str(valor_busqueda)]
        
        if len(df_filtrado) == df_original_len:
            self.logger.warning(f"\n⚠️ {nombre_paso}: No se encontró registro para eliminar en Excel.")
            return False
        
        registros_eliminados = df_original_len - len(df_filtrado)
        
        try:
            df_filtrado.to_excel(file_path, index=False)
            self.logger.info(f"\n✅ {nombre_paso}: {registros_eliminados} registro(s) eliminado(s) y archivo Excel reescrito correctamente.")
            return True
        except Exception as e:
            self.logger.error(f"\n❌ {nombre_paso}: Error al reescribir archivo Excel. Detalles: {e}", exc_info=True)
            return False

    @allure.step("Borrando archivo: '{file_path}'")
    def borrar_archivo(self, file_path: str) -> bool:
        """
        Elimina un archivo del sistema operativo si existe (función de 'eliminar archivo').

        Args:
            file_path (str): Ruta completa al archivo a borrar.

        Returns:
            bool: True si el archivo fue borrado o no existía, False si hubo un error.
        """
        nombre_paso = f"Borrando archivo en: '{file_path}'"
        self.registrar_paso(nombre_paso)
        
        if not os.path.exists(file_path):
            self.logger.info(f"\nℹ️ {nombre_paso}: Archivo no encontrado. No es necesario borrarlo.")
            return True
        
        try:
            os.remove(file_path)
            self.logger.info(f"\n✅ {nombre_paso}: Archivo borrado exitosamente.")
            return True
        except OSError as e:
            self.logger.error(f"\n❌ {nombre_paso}: Error al borrar el archivo. (Permisos, archivo en uso). Detalles: {e}", exc_info=True)
            return False